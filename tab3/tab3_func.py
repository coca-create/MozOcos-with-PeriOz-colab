##Tab3 2つのSRTファイルからデータフレーム表示とExcelファイルを作成します。
import gradio as gr
import pandas as pd
import os
import tempfile
import re
from openpyxl.styles import Alignment, Font, PatternFill

# SRTファイルを解析する関数
def parse_srt(file_path):
    with open(file_path, 'r', encoding='utf-8') as file:
        content = file.read()
    
    pattern = r'(\d+)\n(\d{2}:\d{2}:\d{2},\d{3}) --> (\d{2}:\d{2}:\d{2},\d{3})\n(.*?)\n\n'
    matches = re.findall(pattern, content, re.DOTALL)
    
    subtitles = []
    for match in matches:
        subtitles.append({
            'ID': int(match[0]),
            'Start': match[1],
            'End': match[2],
            'Text': match[3].replace('\n', ' ')
        })
    
    return subtitles

# SRTファイルからExcelファイルを作成する関数
def create_excel_from_srt(english_srt_path=None, japanese_srt_path=None):
    if english_srt_path and japanese_srt_path:
        english_subtitles = parse_srt(english_srt_path)
        japanese_subtitles = parse_srt(japanese_srt_path)

        data = []
        for eng, jap in zip(english_subtitles, japanese_subtitles):
            data.append({
                'ID': eng['ID'],
                'Start': eng['Start'],
                'End': eng['End'],
                'English Subtitle': eng['Text'],
                'Japanese Subtitle': jap['Text']
            })

        df = pd.DataFrame(data)
        base_name = os.path.splitext(os.path.basename(english_srt_path))[0]
        excel_file_name = f"{base_name}.xlsx"
    elif english_srt_path:
        english_subtitles = parse_srt(english_srt_path)

        data = []
        for eng in english_subtitles:
            data.append({
                'ID': eng['ID'],
                'Start': eng['Start'],
                'End': eng['End'],
                'English Subtitle': eng['Text']
            })

        df = pd.DataFrame(data)
        base_name = os.path.splitext(os.path.basename(english_srt_path))[0]
        excel_file_name = f"{base_name}_en.xls"
    elif japanese_srt_path:
        japanese_subtitles = parse_srt(japanese_srt_path)

        data = []
        for jap in japanese_subtitles:
            data.append({
                'ID': jap['ID'],
                'Start': jap['Start'],
                'End': jap['End'],
                'Japanese Subtitle': jap['Text']
            })

        df = pd.DataFrame(data)
        base_name = os.path.splitext(os.path.basename(japanese_srt_path))[0]
        excel_file_name = f"{base_name}_ja.xls"
    else:
        return None, None

    excel_file_path = os.path.join(tempfile.gettempdir(), excel_file_name)
    
    with pd.ExcelWriter(excel_file_path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Subtitles')
        workbook = writer.book
        worksheet = writer.sheets['Subtitles']

        column_widths = {'A': 7, 'B': 25, 'C': 25, 'D': 90, 'E': 90}
        for column, width in column_widths.items():
            worksheet.column_dimensions[column].width = width

        for row in worksheet.iter_rows(min_row=2, max_row=len(df) + 1):
            for cell in row:
                if cell.column_letter == 'A':
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                elif cell.column_letter in ['B', 'C']:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                elif cell.column_letter in ['D', 'E']:
                    cell.alignment = Alignment(horizontal='left', vertical='center')

        for row in worksheet.iter_rows(min_row=2, max_row=len(df) + 1):
            worksheet.row_dimensions[row[0].row].height = 30

        header_font = Font(bold=True)
        for cell in worksheet["1:1"]:
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.fill = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")

    return excel_file_path, df

# コンポーネントの表示を更新してファイルをクリアする関数pd.DataFrame({'1': [''], '2': [''],'3': ['']})
def update_visibility_and_clear(choice):
    clear_update = (gr.update(value=None, visible=True), gr.update(value=None, visible=True), gr.update(value=pd.DataFrame({'1': [''], '2': [''],'3': ['']}), visible=True), gr.update(value=None, visible=True))
    if choice == "only English":
        return (gr.update(value=None, visible=True), gr.update(value=None, visible=False)) + clear_update
    elif choice == "only Japanese":
        return (gr.update(value=None, visible=False), gr.update(value=None, visible=True)) + clear_update
    else:  # "English and Japanese"
        return (gr.update(value=None, visible=True), gr.update(value=None, visible=True)) + clear_update

# ファイルをクリアする関数
def clear_all_files():
    return (gr.update(value=None), gr.update(value=None), gr.update(value=pd.DataFrame({'1': [''], '2': [''],'3': ['']})), gr.update(value=None))

    