import os
import tempfile
import json
from pydub import AudioSegment
import gradio as gr
from faster_whisper import WhisperModel
import zipfile
from docx import Document
import re
from openpyxl.styles import Alignment, Font, PatternFill
import pandas as pd

#tab2の関数
def parse_srt(file_path):
    with open(file_path, 'r', encoding='utf-8') as file:
        content = file.read()
    
    pattern = r'(\d+)\n(\d{2}:\d{2}:\d{2},\d{3}) --> (\d{2}:\d{2}:\d{2},\d{3})\n(.*?)\n\n'
    matches = re.findall(pattern, content, re.DOTALL)
    
    subtitles = []
    for match in matches:
        subtitles.append({
            'ID': int(match[0]),
            'Start': match[1],
            'End': match[2],
            'Text': match[3].replace('\n', ' ')
        })
    
    return subtitles

def create_excel_from_srt(srt_file_path, japanese_srt_path,basename):
    english_subtitles = parse_srt(srt_file_path)
    japanese_subtitles = parse_srt(japanese_srt_path)
    if english_subtitles is None or japanese_subtitles is None:
        return None, pd.DataFrame({'1': [''], '2': [''], '3': ['']})
    else:
        data = []
        for eng, jap in zip(english_subtitles, japanese_subtitles):
            data.append({
                'ID': eng['ID'],
                'Start': eng['Start'],
                'End': eng['End'],
                'English Subtitle': eng['Text'],
                'Japanese Subtitle': jap['Text']
            })

        df = pd.DataFrame(data)
        #base_name = os.path.splitext(os.path.basename(srt_file_path))[0]
        base_name=basename
        excel_file_name = f"{base_name}.xlsx"
        excel_file_path = os.path.join(tempfile.gettempdir(), excel_file_name)
        
        with pd.ExcelWriter(excel_file_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Subtitles')
            workbook = writer.book
            worksheet = writer.sheets['Subtitles']

            column_widths = {'A': 7, 'B': 25, 'C': 25, 'D': 90, 'E': 90}
            for column, width in column_widths.items():
                worksheet.column_dimensions[column].width = width

            for row in worksheet.iter_rows(min_row=2, max_row=len(df) + 1):
                for cell in row:
                    if cell.column_letter == 'A':
                        cell.alignment = Alignment(horizontal='right', vertical='center')
                    elif cell.column_letter in ['B', 'C']:
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    elif cell.column_letter in ['D', 'E']:
                        cell.alignment = Alignment(horizontal='left', vertical='center')

            for row in worksheet.iter_rows(min_row=2, max_row=len(df) + 1):
                worksheet.row_dimensions[row[0].row].height = 30

            header_font = Font(bold=True)
            for cell in worksheet["1:1"]:
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")

        return excel_file_path, df

def count_lines_in_file(file_path):
    try:
        with open(file_path, 'r', encoding='utf-8') as file:
            num_lines = sum(1 for line in file)
        return num_lines
    except FileNotFoundError:
        print(f"Error: File '{file_path}' not found.")
        return 0
    except Exception as e:
        print(f"Error counting lines in file '{file_path}': {str(e)}")
        return 0

def create_translate_files(copied_filename, translate_srt, translate_nr_txt, translate_r_txt, extensions, srt_file_path):
    temp_dir = tempfile.gettempdir()
    output_files = []
    df_data = pd.DataFrame({'1': [''], '2': [''], '3': ['']})

    # エラーチェック
    if not copied_filename:
        error_message = "Error: Please provide a filename."
        dummy_file_path = os.path.join(temp_dir, "error_message.txt")
        with open(dummy_file_path, 'w') as f:
            f.write(error_message)
        return [dummy_file_path], df_data

    translates = {
        "srt": (translate_srt, "_ja.srt"),
        "txt(nr)": (translate_nr_txt, "_NR_ja.txt"),
        "txt(r)": (translate_r_txt, "_R_ja.txt")
    }

    for ext in extensions:
        subtitle_content, extension_suffix = translates.get(ext, ("", ""))
        output_filename = f"{copied_filename}{extension_suffix}"
        output_file_path = os.path.join(temp_dir, output_filename)

        ##ここを追加（翻訳された文字列の編集）
        if ext == 'srt':
            subtitle_content=re.sub(r'[\u200B-\u200D\uFEFF]', '', translate_srt)
            subtitle_content=re.sub(r'\s+', '', subtitle_content)

            pattern = re.compile(r'(\d{1,4})(\d{2}:\d{2}:\d{2},\d{3}-->\d{2}:\d{2}:\d{2},\d{3})')
            matches = pattern.findall(subtitle_content)

            segments = pattern.split(subtitle_content)
            corrected_content = []

            for i in range(1, len(segments), 3):
                segment_id = segments[i]
                timestamp = segments[i + 1]
                text = segments[i + 2]
                
                corrected_content.append(f"{segment_id}")
                corrected_content.append(timestamp.replace('-->', ' --> '))
                corrected_content.append(text)

            # Ensure each subtitle block is separated by exactly one empty line
            final_content = "\n\n".join("\n".join(block) for block in zip(*[iter(corrected_content)]*3))
            with open(output_file_path, 'w', encoding='utf-8') as f:
                f.write(final_content)

            output_files.append(output_file_path)
            continue

        # ここまで

        with open(output_file_path, 'w', encoding='utf-8') as f:
            f.write(subtitle_content)
        
        output_files.append(output_file_path)
    
    
    
    print(f"Output files: {output_files}")

    if "srt" in extensions:
        srt_output_path = next((file for file in output_files if file.endswith("_ja.srt")), None)
        
        if srt_output_path:
            num_lines_english = count_lines_in_file(srt_file_path)
            num_lines_japanese = count_lines_in_file(srt_output_path)
            print(f"num_lines_english: {num_lines_english}, num_lines_japanese: {num_lines_japanese}")

            if num_lines_english == 0 or num_lines_japanese == 0:
                print("Error: One or both of the SRT files are empty or cannot be read.")
                # zipファイルにまとめる。
                if len(output_files)>1:
                    zip_ja_file_name = f"{copied_filename}_ja.zip"
                    zip_ja_file_path = os.path.join(temp_dir, zip_ja_file_name)

                    with zipfile.ZipFile(zip_ja_file_path, 'w') as zip_file:
                        for file in output_files:
                            zip_file.write(file, os.path.basename(file))
                    output_files.append(zip_ja_file_path)        
                return output_files, df_data

            if abs(num_lines_english - num_lines_japanese) > 3:
                print("Error: Number of lines in English and Japanese SRT files do not match.")
                # zipファイルにまとめる。
                if len(output_files)>1:
                    zip_ja_file_name = f"{copied_filename}_ja.zip"
                    zip_ja_file_path = os.path.join(temp_dir, zip_ja_file_name)

                    with zipfile.ZipFile(zip_ja_file_path, 'w') as zip_file:
                        for file in output_files:
                            zip_file.write(file, os.path.basename(file))
                    output_files.append(zip_ja_file_path)        
                return output_files, df_data

            excel_path, df_data = create_excel_from_srt(srt_file_path, srt_output_path,copied_filename)
            print(f"Excel path: {excel_path}, DataFrame: {df_data.head()}")
            if excel_path:
                output_files.append(excel_path)

    # zipファイルにまとめる。
    if len(output_files)>1:
        zip_ja_file_name = f"{copied_filename}_ja.zip"
        zip_ja_file_path = os.path.join(temp_dir, zip_ja_file_name)

        with zipfile.ZipFile(zip_ja_file_path, 'w') as zip_file:
            for file in output_files:
                zip_file.write(file, os.path.basename(file))
        output_files.append(zip_ja_file_path)        
        
    return output_files, df_data



